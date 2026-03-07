# gui.py
# PyQt5 frontend — handles display, user input, and threading.
# scraper.py does the actual fetching; this file just shows the results.

import sys
from datetime import datetime

from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (
    QApplication,
    QCheckBox,
    QDateEdit,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from threading_manager import ThreadingManager
import data_service


APP_STYLE = """
QMainWindow {
    background-color: #c0c0c0;
}

QWidget {
    background-color: #c0c0c0;
    color: #000000;
    font-family: 'Geneva', 'Tahoma', 'Segoe UI', sans-serif;
    font-size: 12px;
}

QLineEdit {
    background-color: #ffffff;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    border-radius: 0px;
    padding: 4px 8px;
    color: #000000;
    font-size: 12px;
    selection-background-color: #6666cc;
    selection-color: #ffffff;
}
QLineEdit:focus {
    border-top: 2px solid #000000;
    border-left: 2px solid #000000;
}

QGroupBox {
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    border-radius: 0px;
    margin-top: 14px;
    padding: 14px 8px 8px 8px;
    font-weight: bold;
    color: #000000;
    font-size: 11px;
    background-color: #c0c0c0;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 4px;
    color: #000000;
}

QPushButton {
    background-color: #c0c0c0;
    border-top: 2px solid #ffffff;
    border-left: 2px solid #ffffff;
    border-right: 2px solid #808080;
    border-bottom: 2px solid #808080;
    border-radius: 0px;
    padding: 5px 14px;
    font-family: 'Geneva', 'Tahoma', sans-serif;
    font-size: 12px;
    color: #000000;
}
QPushButton:hover {
    background-color: #d0d0d0;
}
QPushButton:pressed {
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    background-color: #b0b0b0;
}
QPushButton:disabled {
    color: #808080;
    background-color: #c0c0c0;
}

QPushButton#btn_start {
    background-color: #c0c0c0;
    font-weight: bold;
    color: #000080;
}
QPushButton#btn_stop {
    background-color: #c0c0c0;
    font-weight: bold;
    color: #800000;
}
QPushButton#btn_export_csv,
QPushButton#btn_export_excel {
    background-color: #c0c0c0;
    color: #000000;
}
QPushButton#btn_clear {
    background-color: #c0c0c0;
    color: #404040;
}

QTableWidget {
    background-color: #ffffff;
    alternate-background-color: #eeeeff;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    gridline-color: #d0d0d0;
    color: #000000;
    font-size: 12px;
    selection-background-color: #6666cc;
    selection-color: #ffffff;
}
QTableWidget::item { padding: 3px 6px; }
QTableWidget::item:selected {
    background-color: #6666cc;
    color: #ffffff;
}
QTableWidget::item:alternate { background-color: #eeeeff; }

QHeaderView::section {
    background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #2a8aee, stop:0.4 #1070dd, stop:0.6 #0a5ecc, stop:1 #0848aa);
    color: #ffffff;
    padding: 5px 8px;
    border-top: 1px solid #60aaff;
    border-left: 1px solid #60aaff;
    border-right: 1px solid #0040aa;
    border-bottom: 1px solid #0040aa;
    font-weight: bold;
    font-size: 11px;
}

QProgressBar {
    background-color: #ffffff;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    text-align: center;
    color: #000000;
    font-size: 11px;
    max-height: 14px;
}
QProgressBar::chunk {
    background-color: #6666cc;
    width: 6px;
    margin: 1px;
}

QTextEdit {
    background-color: #ffffff;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    color: #000000;
    font-family: 'Courier New', monospace;
    font-size: 11px;
    padding: 4px;
}

QSpinBox, QDateEdit {
    background-color: #ffffff;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    padding: 4px 8px;
    color: #000000;
}

QCheckBox {
    color: #000000;
    spacing: 6px;
    font-size: 12px;
}
QCheckBox::indicator {
    width: 13px;
    height: 13px;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
    border-right: 2px solid #ffffff;
    border-bottom: 2px solid #ffffff;
    background-color: #ffffff;
    border-radius: 0px;
}
QCheckBox::indicator:checked {
    background-color: #6666cc;
    border-top: 2px solid #808080;
    border-left: 2px solid #808080;
}

QLabel { color: #000000; }

QSplitter::handle {
    background-color: #a0a0a0;
    height: 2px;
}

QScrollBar:vertical {
    background: #c0c0c0;
    width: 16px;
    border: 1px solid #808080;
}
QScrollBar::handle:vertical {
    background: #c0c0c0;
    border-top: 2px solid #ffffff;
    border-left: 2px solid #ffffff;
    border-right: 2px solid #808080;
    border-bottom: 2px solid #808080;
    min-height: 20px;
}
QScrollBar::add-line:vertical,
QScrollBar::sub-line:vertical {
    background: #c0c0c0;
    border-top: 2px solid #ffffff;
    border-left: 2px solid #ffffff;
    border-right: 2px solid #808080;
    border-bottom: 2px solid #808080;
    height: 16px;
}
"""


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("News Scraper — Web Scraping GUI")
        self.setMinimumSize(1150, 780)

        self.all_data = []

        # threading_manager yang handle semua logic scraping di background
        # gui cukup connect ke sinyalnya aja
        self.tm = ThreadingManager()

        self._connect_signals()
        self._init_ui()
        self.setStyleSheet(APP_STYLE)

    def _connect_signals(self):
        # connect ke sinyal dari threading_manager
        # setiap sinyal sudah jalan di main thread jadi aman langsung update widget
        self.tm.update_progress.connect(self._on_progress)
        self.tm.data_ready.connect(self._on_data_ready)
        self.tm.log_message.connect(self._on_log)
        self.tm.error_occurred.connect(self._on_error)

    def _init_ui(self):
        root = QWidget()
        self.setCentralWidget(root)

        outer = QVBoxLayout(root)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        outer.addWidget(self._build_header())

        main = QHBoxLayout()
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)
        main.addWidget(self._build_sidebar())
        main.addWidget(self._build_content())

        main_widget = QWidget()
        main_widget.setLayout(main)
        outer.addWidget(main_widget)

    def _build_header(self):
        widget = QWidget()
        widget.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #000080, stop:0.15 #6666cc,
                stop:0.5 #9999ee, stop:0.85 #6666cc,
                stop:1 #000080);
            border-bottom: 2px solid #000060;
        """)
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(10, 6, 10, 6)

        title = QLabel("News Scraper")
        title.setFont(QFont("Geneva", 12, QFont.Bold))
        title.setStyleSheet("color: #ffffff; background: transparent;")

        self.status_badge = QLabel("● Siap")
        self.status_badge.setStyleSheet("color: #ccffcc; font-weight: bold; font-size: 11px; background: transparent;")

        layout.addWidget(title)
        layout.addStretch()
        layout.addWidget(self.status_badge)
        return widget

    def _build_sidebar(self):
        sidebar = QWidget()
        sidebar.setFixedWidth(280)
        sidebar.setStyleSheet("background-color: #c0c0c0; border-right: 2px solid #808080;")
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(14)

        # URL input
        lbl_url = QLabel("URL Halaman Berita")
        lbl_url.setStyleSheet("color: #000000; font-size: 11px; font-weight: bold;")
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("https://...")
        self.url_input.setMinimumHeight(34)

        # Limit
        lbl_limit = QLabel("Limit Artikel")
        lbl_limit.setStyleSheet("color: #000000; font-size: 11px; font-weight: bold;")
        self.limit_spin = QSpinBox()
        self.limit_spin.setRange(1, 1000)
        self.limit_spin.setValue(20)
        self.limit_spin.setMinimumHeight(32)
        self.limit_spin.setToolTip("Max number of articles to scrape")

        # Keywords — dikirim ke threading_manager buat filter di scraper_core
        lbl_keywords = QLabel("Keywords (opsional)")
        lbl_keywords.setStyleSheet("color: #000000; font-size: 11px; font-weight: bold;")
        self.keywords_input = QLineEdit()
        self.keywords_input.setPlaceholderText("pisahkan dengan koma, misal: politik, ekonomi")
        self.keywords_input.setMinimumHeight(34)

        # Date filter
        self.date_filter_check = QCheckBox("Filter Tanggal")
        self.date_filter_check.setStyleSheet("color: #000000; font-weight: bold;")
        self.date_filter_check.stateChanged.connect(self._toggle_date_filter)

        lbl_from = QLabel("Dari")
        lbl_from.setStyleSheet("color: #000000; font-size: 11px;")
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_from.setEnabled(False)
        self.date_from.setMinimumHeight(32)

        lbl_to = QLabel("Sampai")
        lbl_to.setStyleSheet("color: #000000; font-size: 11px;")
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setEnabled(False)
        self.date_to.setMinimumHeight(32)

        # Buttons
        self.btn_start = QPushButton("Mulai Scraping")
        self.btn_start.setObjectName("btn_start")
        self.btn_start.setMinimumHeight(36)
        self.btn_start.clicked.connect(self._start_scraping)

        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setObjectName("btn_stop")
        self.btn_stop.setMinimumHeight(36)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._request_stop)

        self.btn_export_csv = QPushButton("Export CSV")
        self.btn_export_csv.setObjectName("btn_export_csv")
        self.btn_export_csv.setMinimumHeight(34)
        self.btn_export_csv.setEnabled(False)
        self.btn_export_csv.clicked.connect(lambda: self._export_data("csv"))

        self.btn_export_excel = QPushButton("Export Excel")
        self.btn_export_excel.setObjectName("btn_export_excel")
        self.btn_export_excel.setMinimumHeight(34)
        self.btn_export_excel.setEnabled(False)
        self.btn_export_excel.clicked.connect(lambda: self._export_data("excel"))

        self.btn_clear = QPushButton("Clear")
        self.btn_clear.setObjectName("btn_clear")
        self.btn_clear.setMinimumHeight(34)
        self.btn_clear.clicked.connect(self._clear_all)

        # divider helper
        def divider():
            d = QWidget()
            d.setFixedHeight(1)
            d.setStyleSheet("background-color: #808080;")
            return d

        layout.addWidget(lbl_url)
        layout.addWidget(self.url_input)
        layout.addWidget(divider())
        layout.addWidget(lbl_limit)
        layout.addWidget(self.limit_spin)
        layout.addWidget(lbl_keywords)
        layout.addWidget(self.keywords_input)
        layout.addWidget(divider())
        layout.addWidget(self.date_filter_check)
        layout.addWidget(lbl_from)
        layout.addWidget(self.date_from)
        layout.addWidget(lbl_to)
        layout.addWidget(self.date_to)
        layout.addWidget(divider())
        layout.addWidget(self.btn_start)
        layout.addWidget(self.btn_stop)
        layout.addSpacing(4)
        layout.addWidget(self.btn_export_csv)
        layout.addWidget(self.btn_export_excel)
        layout.addWidget(self.btn_clear)
        layout.addStretch()

        return sidebar

    def _build_content(self):
        content = QWidget()
        content.setStyleSheet("background-color: #c0c0c0;")
        layout = QVBoxLayout(content)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        # progress row
        prog_widget = QWidget()
        prog_widget.setStyleSheet("background: transparent;")
        prog_layout = QHBoxLayout(prog_widget)
        prog_layout.setContentsMargins(0, 0, 0, 0)
        prog_layout.setSpacing(10)

        self.progress_label = QLabel("Siap.")
        self.progress_label.setFixedWidth(200)
        self.progress_label.setStyleSheet("color: #000000; font-size: 12px;")

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setMinimumHeight(6)
        self.progress_bar.setMaximumHeight(6)
        self.progress_bar.setFormat("")

        prog_layout.addWidget(self.progress_label)
        prog_layout.addWidget(self.progress_bar)

        # table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["No", "Judul Berita", "Tanggal", "Isi (Preview)", "URL"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 45)
        self.table.setColumnWidth(2, 130)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setMinimumHeight(260)

        self.total_label = QLabel("Total: 0 artikel ditemukan")
        self.total_label.setStyleSheet("color: #000000; font-size: 12px;")

        # log
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(130)

        layout.addWidget(prog_widget)
        layout.addWidget(self.table)
        layout.addWidget(self.total_label)
        lbl_log = QLabel("Log Aktivitas")
        lbl_log.setStyleSheet("color: #000000; font-weight: bold; font-size: 11px;")
        layout.addWidget(lbl_log)
        layout.addWidget(self.log_box)

        return content

    # -------------------------------------------------------------------------

    def _toggle_date_filter(self, state):
        enabled = state == Qt.Checked
        self.date_from.setEnabled(enabled)
        self.date_to.setEnabled(enabled)

    def _start_scraping(self):
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "URL Kosong", "Masukkan URL halaman berita terlebih dahulu!")
            return
        if not url.startswith("http"):
            QMessageBox.warning(self, "URL Tidak Valid", "URL harus diawali dengan http:// atau https://")
            return

        # bersihkan hasil scraping sebelumnya sebelum mulai yang baru
        self.all_data.clear()
        self.table.setRowCount(0)
        self.total_label.setText("Total: 0 artikel ditemukan")
        self.progress_bar.setValue(0)

        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_export_csv.setEnabled(False)
        self.btn_export_excel.setEnabled(False)
        self.status_badge.setText("Scraping...")
        self.status_badge.setStyleSheet("color: #ffff99; font-weight: bold; font-size: 11px; background: transparent;")

        # parse keywords dari input — pisah berdasarkan koma, buang spasi
        raw_keywords = self.keywords_input.text().strip()
        keywords = [k.strip() for k in raw_keywords.split(",") if k.strip()] if raw_keywords else []

        self._log(f"Mulai scraping: {url}")
        if keywords:
            self._log(f"Keywords: {', '.join(keywords)}")
        if self.date_filter_check.isChecked():
            self._log(
                f"Filter tanggal: {self.date_from.date().toString('dd-MM-yyyy')} "
                f"s/d {self.date_to.date().toString('dd-MM-yyyy')}"
            )

        # lempar juga limit dan stop_flag biar threading_manager bisa handle stop & limit
        self._stop_flag[0] = False  # reset dulu setiap kali mulai scraping baru
        limit_artikel = self.limit_spin.value()
        self.tm.start_scraping_task(url, keywords, limit_artikel, self._stop_flag)

    # dibungkus list biar bisa di-mutate dari thread lain
    # kalau pakai bool biasa, perubahan nilai di thread tidak kelihatan dari thread lain
    @property
    def _stop_flag(self):
        if not hasattr(self, "_stop_flag_list"):
            self._stop_flag_list = [False]
        return self._stop_flag_list

    def _request_stop(self):
        self._stop_flag[0] = True
        self.btn_stop.setEnabled(False)
        self._log("Stop diminta — menunggu artikel saat ini selesai...")

    # -------------------------------------------------------------------------

    def _on_progress(self, value: int):
        self.progress_bar.setValue(value)
        self.progress_label.setText(f"Memproses... {value}%")
        if value == 100:
            self.progress_label.setText("Selesai!")

    def _on_data_ready(self, data: list):
        # data_ready ngirim semua hasil sekaligus dalam bentuk list
        # loop di sini buat masukin satu-satu ke tabel
        for item in data:
            row = self.table.rowCount()
            self.table.insertRow(row)

            no_item = QTableWidgetItem(str(row + 1))
            no_item.setTextAlignment(Qt.AlignCenter)

            judul   = item.get("judul",   "-")
            tanggal = item.get("tanggal", "-")
            isi     = item.get("isi",     "-")
            url     = item.get("url",     "-")
            preview = (isi[:120] + "...") if len(isi) > 120 else isi

            self.table.setItem(row, 0, no_item)
            self.table.setItem(row, 1, QTableWidgetItem(judul))
            self.table.setItem(row, 2, QTableWidgetItem(tanggal))
            self.table.setItem(row, 3, QTableWidgetItem(preview))
            self.table.setItem(row, 4, QTableWidgetItem(url))

            self.all_data.append({"judul": judul, "tanggal": tanggal, "isi": isi, "url": url})

        self.total_label.setText(f"Total: {len(self.all_data)} artikel ditemukan")
        self.table.scrollToBottom()
        self._on_finished()

    def _on_log(self, msg: str):
        self._log(msg)

    def _on_finished(self):
        self._reset_ui_after_scraping()
        self._log(f"Selesai. Total artikel: {len(self.all_data)}")
        self.progress_bar.setValue(100)
        self.progress_label.setText(f"Selesai - {len(self.all_data)} artikel")
        if self.all_data:
            self.btn_export_csv.setEnabled(True)
            self.btn_export_excel.setEnabled(True)

    def _on_error(self, msg: str):
        self._log(f"ERROR: {msg}")
        QMessageBox.critical(self, "Terjadi Error", f"Scraping gagal:\n\n{msg}")
        self._reset_ui_after_scraping()

    # -------------------------------------------------------------------------

    def _export_data(self, fmt: str):
        if not self.all_data:
            QMessageBox.information(self, "Tidak Ada Data", "Belum ada data untuk diexport.")
            return

        df = pd.DataFrame(self.all_data, columns=["judul", "tanggal", "isi", "url"])
        df.index += 1

        # article body has \n\n between paragraphs — collapses into one cell cleanly
        df["isi"] = df["isi"].str.replace(r"\n+", " ", regex=True).str.strip()

        if fmt == "csv":
            path, _ = QFileDialog.getSaveFileName(
                self, "Simpan sebagai CSV", "hasil_scraping.csv", "CSV Files (*.csv)"
            )
            if path:
                df.to_csv(path, index_label="No", encoding="utf-8-sig")
                self._log(f"Saved: {path}")
                QMessageBox.information(self, "Sukses", f"File CSV berhasil disimpan!\n{path}")

        elif fmt == "excel":
            path, _ = QFileDialog.getSaveFileName(
                self, "Simpan sebagai Excel", "hasil_scraping.xlsx", "Excel Files (*.xlsx)"
            )
            if path:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                df.to_excel(path, index_label="No")

                # pandas doesn't support styling, so we reopen with openpyxl to apply it
                wb = load_workbook(path)
                ws = wb.active

                # make headers visually distinct so readers don't mistake them for data rows
                header_fill = PatternFill("solid", fgColor="1565C0")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # without this, headers scroll away and 500-row exports become unreadable
                ws.freeze_panes = "A2"

                # isi column intentionally wide — article bodies are long
                ws.column_dimensions["A"].width = 5   # No
                ws.column_dimensions["B"].width = 8   # index
                ws.column_dimensions["C"].width = 55  # judul
                ws.column_dimensions["D"].width = 25  # tanggal
                ws.column_dimensions["E"].width = 80  # isi
                ws.column_dimensions["F"].width = 50  # url

                # row height is estimated from char count — openpyxl can't measure actual rendered height
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    # kolom urutan: No(A), index(B), judul(C), tanggal(D), isi(E), url(F)
                    # isi ada di kolom 5 (E) — sesuai output pandas
                    isi_val = str(ws.cell(row=row[0].row, column=5).value or "")
                    estimated_lines = max(1, len(isi_val) // 80)
                    ws.row_dimensions[row[0].row].height = min(estimated_lines * 15, 150)

                wb.save(path)
                self._log(f"Saved: {path}")
                QMessageBox.information(self, "Sukses", f"File Excel berhasil disimpan!\n{path}")

    def _clear_all(self):
        self.table.setRowCount(0)
        self.all_data.clear()
        self.log_box.clear()
        self.progress_bar.setValue(0)
        self.progress_label.setText("Siap.")
        self.total_label.setText("Total: 0 artikel ditemukan")
        self.btn_export_csv.setEnabled(False)
        self.btn_export_excel.setEnabled(False)
        self.status_badge.setText("Siap")
        self.status_badge.setStyleSheet("color: #ccffcc; font-weight: bold; font-size: 11px; background: transparent;")

    def _reset_ui_after_scraping(self):
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._stop_flag[0] = False
        self.status_badge.setText("Siap")
        self.status_badge.setStyleSheet("color: #ccffcc; font-weight: bold; font-size: 11px; background: transparent;")

    def _log(self, msg: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_box.append(f"[{timestamp}]  {msg}")


# if __name__ == "__main__":
#     app = QApplication(sys.argv)
#     app.setStyle("Fusion")
#     window = MainWindow()
#     window.show()
#     sys.exit(app.exec_())