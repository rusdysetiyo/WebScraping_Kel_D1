# data_service.py
import csv
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_to_csv(data, path):
    try:
        if not data:
            return False

        keys = data[0].keys()

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(data)

        return True

    except Exception as e:
        print("CSV Export Error:", e)
        return False


def export_to_excel(data, path):
    try:
        if not data:
            return False

        df = pd.DataFrame(data)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data")

            worksheet = writer.sheets["Data"]

            # STYLE HEADER
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")

            header_alignment = Alignment(horizontal="center", vertical="center")

            cell_alignment = Alignment(vertical="center", wrap_text=True)

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # HEADER FORMAT
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # CELL FORMAT
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    cell.border = border
                    cell.alignment = cell_alignment

            # AUTO WIDTH COLUMN
            for column in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

            # FREEZE HEADER
            worksheet.freeze_panes = "A2"

            # FILTER
            worksheet.auto_filter.ref = worksheet.dimensions

            # EXCEL TABLE STYLE
            table = Table(
                displayName="DataTable",
                ref=worksheet.dimensions
            )

            style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )

            table.tableStyleInfo = style
            worksheet.add_table(table)

        return True

    except Exception as e:
        print("Excel Export Error:", e)
        return False